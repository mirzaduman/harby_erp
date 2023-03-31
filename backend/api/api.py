import base64
import uuid
from datetime import timedelta, datetime

import openpyxl
from django.contrib.auth.password_validation import validate_password
from django.core.files.base import ContentFile
from django.template.loader import render_to_string
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from pytz import timezone
from random import randint
from typing import List
from decimal import Decimal
import json
from django.shortcuts import get_object_or_404
from django.http import Http404, HttpResponse
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.db.models import Sum

from ninja import NinjaAPI, Form, File, Query
from ninja.files import UploadedFile
from weasyprint import HTML

from .schemas import CreateTourSchema, TourDaySchema, TrucksSchema, GetTourSchema, GetToursSchema, UpdateTourSchema, \
    GetToursFilterSchema, CreateTruckSchema, AddTruckDocumentSchema, ChangeTruckPayment, CreateContactSchema, \
    UpdateContactSchema, CreateMeetingSchema, CreateFuelCardSchema, UpdateFuelcardSchema, ChangeFuelCardDriverSchema, \
    GetContactsSchema, GetFuelCardsSchema, GetBillsSchema, CreateBillSchema, ProductSchema, GetGutschrifts, \
    CreateGutschrift, UpdateGutschrift, CreateGutschriftPayment, CreateWorker, UpdateWorker, AddDebt, AddWorkerDocument, \
    WorkTimeSchema, CreateOffdays, UpdateOffdays, CreateTag, UpdateTag, CreateCustomer, UpdateCustomer, CreateOwnFirm, \
    UpdateOwnFirm, CreateAdmin, UpdateAdmin, DeleteTour, GetLogs, Login, DeleteTag, GetAdminDetails, GetOwnFirm, \
    GetContact, GetFuelcard, GetGutschrift, GetWorkers, GetDailyExpenses, GetCustomers, GetAdmins, GetOffdays, \
    UpdateMeeting, DeleteCustomer, GetHomepage, UpdateTruck, UpdateTruckDocument, UpdateDebt, UpdateWorkerDocument

from backend.trucks_db import Truck, TruckDocument
from backend.workers_db import Worker, Offday, OffdayTag, Position, DebtPayment, WorkerDocument, WorkTime, \
    HolidayAccount
from backend.tours_db import Tour, TourDay, TourStatus
from backend.firms_db import Firm
from backend.own_firms_db import OwnFirm
from backend.contacts_db import Contact, ContactTag, Meeting
from backend.fuel_cards import Fuelcard, FuelcardActivities, FuelcardFirm
from backend.bills_db import Bill, Product
from backend.gutschriften_db import Gutschrift, GutschriftPayment
from backend.settings_db import Colour, HarbyAdmin, Log

import pandas as pd

api = NinjaAPI()

today = datetime.now(timezone('Europe/Berlin')).date()

root = 'https://elbcargo-server.harby.de'


def format_number(number, precision=2):
    # build format string
    format_str = '{{:,.{}f}}'.format(precision)

    # make number string
    number_str = format_str.format(number)

    # replace chars
    return number_str.replace(',', 'X').replace('.', ',').replace('X', '.')


def if_null(obj):
    if obj:
        return obj
    else:
        return ''


# SHARED
@api.get('/get-firms', tags=['Global'])
def get_firms(request, own_firm: str = Query(...)):
    def own_firm_query():
        own_firm_data = get_object_or_404(OwnFirm, name=own_firm)
        return own_firm_data

    tour_firms = []
    contacts_firms = []
    fuelcard_firms = []
    own_firm_q = get_object_or_404(OwnFirm, name=own_firm)
    for tour_firm in Firm.objects.filter(own_firm=own_firm_q):
        tour_firms.append(
            {'id': tour_firm.id, 'name': tour_firm.name, 'vat': tour_firm.vat, 'address': tour_firm.address})
    for contacts_firm in Contact.objects.filter(own_firm=own_firm_q, firm__isnull=False).values('firm').distinct():
        contacts_firms.append(contacts_firm['firm'])
    for fuelcard_firm in FuelcardFirm.objects.all():
        fuelcard_firms.append({'id': fuelcard_firm.id, 'name': fuelcard_firm.name})
    return {'tour_firms': tour_firms, 'contacts_firms': contacts_firms, 'fuelcard_firms': fuelcard_firms}


@api.get('/get-drivers', tags=['Global'])
def get_drivers(request, own_firm: str = Query(...)):
    drivers = Worker.objects.filter(own_firm=get_object_or_404(OwnFirm, name=own_firm), is_driver=True)
    response = []
    for driver in drivers:
        driver_details = {'id': driver.id, 'name': driver.name}
        response.append(driver_details)
    return response


# HOMEPAGE
@api.get('/get-homepage', tags=['Homepage'])
def get_homepage(request, query: GetHomepage = Query(...)):
    own_firm = get_object_or_404(OwnFirm, name=query.own_firm)
    response = {}
    last_12_months = []
    for i in range(12):
        month = today.month - i
        year = today.year
        if month > 0:
            last_12_months.append({'month': month, 'year': year})
        else:
            last_12_months.append({'month': 12 + month, 'year': year - 1})
    sales = {}
    total = 0
    sales_months = []
    for i in last_12_months:
        bill_sales = Bill.objects.filter(own_firm=own_firm, creation_date__month=i['month'],
                                         creation_date__year=i['year']).aggregate(Sum('end_sum'))
        bill = bill_sales['end_sum__sum']
        if not bill:
            bill = 0
        else:
            try:
                bill = Decimal(round(bill, 2))
            except TypeError:
                bill = 0
        gutschrift_sales = GutschriftPayment.objects.filter(gutschrift__own_firm=own_firm, date__month=i['month'],
                                                            date__year=i['year']).aggregate(
            Sum('amount'))
        gutschrift = gutschrift_sales['amount__sum']
        if not gutschrift:
            gutschrift = 0
        else:
            try:
                gutschrift = Decimal(round(gutschrift, 2))
            except TypeError:
                gutschrift = 0
        sales_months.append({'month': i['month'], 'year': i['year'], 'bills': bill, 'gutschrifts': gutschrift,
                             'total': bill + gutschrift})
        total += bill + gutschrift
    sales['total'] = total
    sales['months'] = sales_months
    response['sales'] = sales
    tour_days = []
    for tour_day in TourDay.objects.filter(tour__own_firm=own_firm, date=today):
        drivers = []
        for driver in tour_day.drivers.all():
            drivers.append({'id': driver.id, 'name': driver.name})
        tour_obj = tour_day.tour

        def get_plate():
            if tour_day.vehicle:
                return tour_day.vehicle.plate
            else:
                return None

        tour_days.append(
            {'id': tour_obj.id, 'roller_nr': tour_obj.roller_nr, 'plate': get_plate(), 'drivers': drivers,
             'firm': tour_obj.firm.name})
    response['tours'] = tour_days
    appointments = {}
    vehicle_appointments = []
    driver_appointments = []
    appointments['vehicles'] = vehicle_appointments
    appointments['drivers'] = driver_appointments
    for t_d in TruckDocument.objects.filter(truck__own_firm=own_firm, expiry_date__isnull=False, done=False).order_by(
            'expiry_date')[:8]:
        vehicle_appointments.append(
            {'id': t_d.id, 'plate': t_d.truck.plate, 'name': t_d.name,
             'expiry_date': t_d.expiry_date.strftime("%d.%m.%Y")})
    for w_d in WorkerDocument.objects.filter(worker__own_firm=own_firm, expiry_date__isnull=False, done=False).order_by(
            'expiry_date')[:8]:
        driver_appointments.append(
            {'id': w_d.id, 'worker_id': w_d.worker.id, 'driver': w_d.worker.name, 'name': w_d.name,
             'expiry_date': w_d.expiry_date.strftime("%d.%m.%Y")})
    response['appointments'] = appointments
    payments = {}
    gutschrift_q = \
        GutschriftPayment.objects.filter(gutschrift__own_firm=own_firm, date__month=today.month).aggregate(
            Sum('amount'))
    if gutschrift_q:
        gutschrift = gutschrift_q['amount__sum']
        try:
            gutschrift = Decimal(round(gutschrift, 2))
        except TypeError:
            gutschrift = 0
    else:
        gutschrift = 0
    payments['gutschrift_total'] = gutschrift
    bill_q = Bill.objects.filter(own_firm=own_firm, creation_date__month=today.month).aggregate(Sum('end_sum'))
    if bill_q:
        bill = bill_q['end_sum__sum']
        try:
            bill = Decimal(round(bill, 2))
        except TypeError:
            bill = 0
    else:
        bill = 0
    payments['bills_total'] = bill
    response['payments'] = payments
    return response


# TOUR
@api.get('/get-tour', tags=['Tour'])
def get_tour(request, query: GetTourSchema = Query(...)):
    tour = Tour.objects.get(roller_nr=query.roller_nr, own_firm=get_object_or_404(OwnFirm, name=query.own_firm))
    tour_details = {'id': tour.id, 'roller_nr': tour.roller_nr, 'general_notes': tour.general_notes,
                    'firm': str(tour.firm), 'default_truck': str(tour.default_truck),
                    'default_driver': str(tour.default_driver), 'default_worker_id': str(tour.default_driver)}
    tour_day_list = []
    if query.period_month and query.period_month:
        queryset = TourDay.objects.filter(tour=tour, date__year=query.period_year,
                                          date__month=query.period_month)
    else:
        queryset = TourDay.objects.filter(tour=tour)
    for tour_day in queryset:
        tour_day_details = {'id': tour_day.id, 'date': str(tour_day.date), 'status': str(tour_day.status)}
        tour_day_drivers = []
        for tour_day_driver in tour_day.drivers.all():
            tour_day_drivers.append(str(tour_day_driver))
        tour_day_details['drivers'] = tour_day_drivers
        tour_day_details['daily_note'] = tour_day.daily_note
        tour_day_details['vehicle'] = str(tour_day.vehicle)
        tour_day_list.append(tour_day_details)
    tour_details['tour_days'] = tour_day_list
    return tour_details


@api.get('/get-tours', tags=['Tour'])
def get_tours(request, query: GetToursSchema = Query(...)):
    own_firm = get_object_or_404(OwnFirm, name=query.own_firm)
    tours = Tour.objects.filter(own_firm=own_firm)
    if query.firm:
        firm = get_object_or_404(Firm, name=query.firm)
        tours = tours.filter(firm=firm)
    tours_response = []
    for tour in tours:
        qs = TourDay.objects.filter(tour=tour, date__year=query.year, date__month=query.month).order_by('date')
        if query.status:
            status = get_object_or_404(TourStatus, name=query.status)
            qs = qs.filter(status=status)
        if query.drivers:
            driver_list = []
            driver_query_list = [i.split(',') for i in query.drivers]
            for driver in driver_query_list[0]:
                driver_query = get_object_or_404(Worker, id=driver)
                driver_list.append(driver_query)
            qs = qs.filter(drivers__in=driver_list)
        tour_days = []
        for q in qs:
            tour_day_drivers = []
            for tour_day_driver in q.drivers.all():
                tour_day_drivers.append({'id': tour_day_driver.id, 'name': tour_day_driver.name})
            tour_day_details = {'id': q.id, 'date': str(q.date), 'status': str(q.status), 'drivers': tour_day_drivers,
                                'daily_note': q.daily_note, 'truck': str(q.vehicle)}
            tour_days.append(tour_day_details)

        def get_firm():
            if tour.firm:
                return tour.firm.name
            else:
                return None

        def get_default_driver():
            if tour.default_driver:
                return {'id': tour.default_driver.id, 'name': tour.default_driver.name}
            else:
                return None

        def get_default_truck():
            if tour.default_truck:
                return {'id': tour.default_truck.id, 'name': tour.default_truck.plate}
            else:
                return None

        tour_details = {'roller_nr': tour.roller_nr, 'default_driver': get_default_driver(),
                        'default_truck': get_default_truck(), 'general_notes': tour.general_notes, 'firm': get_firm(),
                        'tour_days': tour_days}
        tours_response.append(tour_details)
    response = []
    for t in tours_response:
        if any(t['firm'] in d for d in response):
            for i in response:
                if t['firm'] in i:
                    i[t['firm']].append(t)
        else:
            response.append({t['firm']: [t]})
    return response


@api.post('/create-tour', tags=['Tour'])
def create_tour(request, data: CreateTourSchema = Form(...)):
    def firm_finder():
        firm = get_object_or_404(Firm, name=data.firm_name)
        return firm

    def truck_finder():
        if data.default_truck:
            truck = get_object_or_404(Truck, plate=data.default_truck)
        else:
            truck = None
        return truck

    def driver_finder():
        if data.default_driver:
            driver = get_object_or_404(Worker, id=data.default_driver)
        else:
            driver = None
        return driver

    try:
        created_tour = Tour.objects.create(roller_nr=data.roller_nr, firm=firm_finder(),
                                           general_notes=data.general_note,
                                           own_firm=get_object_or_404(OwnFirm, name=data.own_firm),
                                           default_truck=truck_finder(), default_driver=driver_finder())
    except IntegrityError:
        return HttpResponse('Eine Tour mit dieser Roller Nr. existiert bereits.', status=406)

    if created_tour:
        log_input = f'Tour {created_tour.roller_nr} wurde erstellt.'
        Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=created_tour.own_firm,
                           log_input=log_input)
    return 200


@api.post('/update-tour', tags=['Tour'])
def update_tour(request, data: UpdateTourSchema = Form(...)):
    tour = get_object_or_404(Tour, roller_nr=data.query_roller_nr,
                             own_firm=get_object_or_404(OwnFirm, name=data.own_firm))
    old_tour = tour
    log_changes = ''
    if data.roller_nr:
        if Tour.objects.filter(roller_nr=data.roller_nr).exists():
            return HttpResponse('Eine Tour mit dieser Roller Nr. existiert bereits.', status=406)
        else:
            log_changes += f'Roller Nr.: {old_tour.roller_nr} => {data.roller_nr}; '
            tour.roller_nr = data.roller_nr
    if data.firm_name:
        firm = get_object_or_404(Firm, name=data.firm_name)
        log_changes += f'Firma: {old_tour.firm} => {firm}; '
        tour.firm = firm
    if data.general_note:
        log_changes += f'Notizen: {if_null(old_tour.general_notes)} => {data.general_note}; '
        tour.general_notes = data.general_note
    if data.default_driver:
        driver = get_object_or_404(Worker, id=data.default_driver)
        log_changes += f'Fahrer: {if_null(old_tour.default_driver)} => {driver.name}; '
        tour.default_driver = driver
    if data.default_truck:
        truck = get_object_or_404(Truck, plate=data.default_truck)
        log_changes += f'LKW: {if_null(old_tour.default_truck)} => {truck}; '
        tour.default_truck = truck
    tour.save()
    log_input = f'Tour {old_tour.roller_nr} wurde verändert. Veränderungen: {log_changes}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=tour.own_firm,
                       log_input=log_input)
    return 200


@api.post('/delete-tour', tags=['Tour'])
def delete_tour(request, data: DeleteTour = Form(...)):
    def own_firm_query():
        own_firm_data = get_object_or_404(OwnFirm, name=data.own_firm)
        return own_firm_data

    tour = get_object_or_404(Tour, roller_nr=data.roller_nr, own_firm=get_object_or_404(OwnFirm, name=data.own_firm))
    tour_nr = tour.roller_nr
    tour.delete()
    log_input = f'Tour {tour_nr} wurde gelöscht.'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=tour.own_firm,
                       log_input=log_input)
    return 200


@api.post('/create-tour-days', tags=['Tour'])
def create_tour_days(request, data: TourDaySchema = Form(...)):
    tour = get_object_or_404(Tour, own_firm=get_object_or_404(OwnFirm, name=data.own_firm), roller_nr=data.roller_nr)

    def get_driver_objects():
        drivers = []
        driver_query_list = [i.split(',') for i in data.driver_list]
        for i in driver_query_list[0]:
            drivers.append(i)
        return drivers

    dates_list = [i.split(',') for i in data.dates]
    created_tour_days = []
    created_days_string = ''
    for date in dates_list[0]:
        if TourDay.objects.filter(tour=tour, date=date).exists():
            for tour_day in TourDay.objects.filter(tour=tour, date=date):
                tour_day.delete()
        date_object = datetime.strptime(date, '%Y-%m-%d')
        if date_object.weekday() < 5:
            created_tour_day = TourDay.objects.create(date=date, status=get_object_or_404(TourStatus, name=data.status),
                                                      daily_note=data.note,
                                                      vehicle=get_object_or_404(Truck, plate=data.vehicle), tour=tour)
            for driver in get_driver_objects():
                created_tour_day.drivers.add(Worker.objects.get(id=driver))
            created_tour_days.append(created_tour_day.date)
            created_days_string += f'{datetime.strftime(date_object, "%d.%m.%Y")}, '
    if created_tour_days:
        log_input = f'Neue Tour Tage wurden für die Tour {tour.roller_nr} hinzugefügt. Tage: {created_days_string}'
        Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=tour.own_firm,
                           log_input=log_input)
        return 200
    else:
        return ''


@api.post('/delete-tour-days', tags=['Tour'])
def delete_tour_days(request, tour_days: List[str] = Form(...), admin: str = Form(...)):
    dates_list = [i.split(',') for i in tour_days]
    log_input_string = ''
    n = 1
    own_firm = None
    tour_nr = ''
    for tour_day in dates_list[0]:
        tour_day = get_object_or_404(TourDay, id=tour_day)
        if n == 1:
            own_firm = tour_day.tour.own_firm
            tour_nr = tour_day.tour.roller_nr
            n = 0
        log_input_string += f'{datetime.strftime(tour_day.date, "%d.%m.%Y")}, '
        tour_day.delete()
    log_input = f'Tour Tage von der Tour {tour_nr} wurden gelöscht. Gelöschte Tage: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=own_firm, log_input=log_input)
    return 200


# TRUCK
@api.get('/get-trucks', tags=['Truck'], description='order_by options: plate, manufacturer')
def get_trucks(request, query: TrucksSchema = Query(...)):
    if query.order_by:
        if query.order_by == 'plate':
            if query.direction == 'asc':
                qs = Truck.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('plate')
            elif query.direction == 'des':
                qs = Truck.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-plate')
        elif query.order_by == 'manufacturer':
            if query.direction == 'asc':
                qs = Truck.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'manufacturer')
            elif query.direction == 'des':
                qs = Truck.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-manufacturer')
    else:
        qs = Truck.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm))
    if query.plate:
        qs = qs.filter(plate__contains=query.plate)
    if query.manufacturer:
        qs = qs.filter(manufacturer__contains=query.manufacturer)
    if query.model:
        qs = qs.filter(model__contains=query.model)
    trucks = []
    if qs:
        for truck in qs:
            truck_details = {
                'id': truck.id,
                'plate': truck.plate,
                'manufacturer': truck.manufacturer,
                'model': truck.model
            }
            trucks.append(truck_details)
    return trucks


@api.post('/create-truck', tags=['Truck'])
def create_truck(request, data: CreateTruckSchema = Form(...)):
    own_firm = get_object_or_404(OwnFirm, name=data.own_firm)
    try:
        created_truck = Truck.objects.create(own_firm=own_firm, plate=data.plate,
                                             manufacturer=data.manufacturer,
                                             model=data.model)
    except IntegrityError:
        return HttpResponse('Es existiert bereits ein LKW mit diesem Kennzeichen.', status=406)
    log_input = f'LKW {created_truck.plate} wurde hinzugefügt.'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=own_firm,
                       log_input=log_input)
    return 200


@api.get('/get-truck', tags=['Truck'])
def get_truck(request, own_firm: str, plate: str):
    truck = get_object_or_404(Truck, own_firm=get_object_or_404(OwnFirm, name=own_firm), plate=plate)
    truck_details = {'own_firm': str(truck.own_firm), 'id': truck.id, 'plate': truck.plate,
                     'manufacturer': truck.manufacturer,
                     'model': truck.model, 'price': truck.price, 'payment_method': truck.payment_method,
                     'paid_day': truck.paid_day, 'paid_status': truck.paid_status,
                     'total_installment_months': truck.total_installment_months,
                     'installment_monthly_payment_amount': truck.installment_monthly_payment_amount,
                     'installment_start_date': truck.installment_start_date,
                     'installment_end_date': truck.installment_end_date}
    truck_documents = []
    for document in TruckDocument.objects.filter(truck=truck).order_by('-id'):
        document_details = {'id': document.id, 'name': document.name, 'expiry_date': document.expiry_date,
                            'done': document.done}
        if document.file:
            document_details['file_url'] = root + document.file.url
        else:
            document_details['file_url'] = None
        truck_documents.append(document_details)
    return {'details': truck_details, 'documents': truck_documents}


@api.post('/delete-truck', tags=['Truck'])
def delete_truck(request, own_firm: str = Form(...), plate: str = Form(...), admin: str = Form(...)):
    own_firm_q = get_object_or_404(OwnFirm, name=own_firm)
    truck = get_object_or_404(Truck, own_firm=own_firm_q, plate=plate)
    truck_plate = truck.plate
    truck.delete()
    log_input = f'LKW {truck_plate} wurde gelöscht.'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=own_firm_q, log_input=log_input)
    return 200


@api.post('/update-truck', tags=['Truck'])
def update_truck(request, data: UpdateTruck = Form(...)):
    own_firm_q = get_object_or_404(OwnFirm, name=data.own_firm)
    truck = get_object_or_404(Truck, own_firm=own_firm_q, id=data.truck_id)
    log_input_string = ''
    if data.plate:
        log_input_string += f'Kennzeichen: {truck.plate} => {data.plate}; '
        truck.plate = data.plate
    if data.manufacturer:
        log_input_string += f'Marke: {truck.manufacturer} => {data.manufacturer}; '
        truck.manufacturer = data.manufacturer
    if data.model:
        log_input_string += f'Modell: {truck.model} => {data.model}; '
        truck.model = data.model
    truck.save()
    log_input = f'Daten zum LKW {truck.plate} wurden verändert. Veränderungen: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=own_firm_q,
                       log_input=log_input)
    return 200


@api.post('/add-truck-document', tags=['Truck'])
def add_truck_document(request, data: AddTruckDocumentSchema = Form(...)):
    own_firm = get_object_or_404(OwnFirm, name=data.own_firm)

    def date_query():
        if data.expiry_date:
            return data.expiry_date
        else:
            return None

    created_truck_document = TruckDocument.objects.create(
        truck=get_object_or_404(Truck, plate=data.plate, own_firm=own_firm), name=data.name, expiry_date=date_query())
    if data.done:
        created_truck_document.done = True
        created_truck_document.save()
    log_input = f'Neues Dokument mit dem Namen {created_truck_document.name} wurde für das LKW {created_truck_document.truck.plate} hinzugefügt.'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=own_firm,
                       log_input=log_input)
    return created_truck_document.id


@api.post('/update-truck-document', tags=['Truck'])
def update_truck_document(request, data: UpdateTruckDocument = Form(...)):
    own_firm = get_object_or_404(OwnFirm, name=data.own_firm)
    document = get_object_or_404(TruckDocument, id=data.document_id)
    log_input_string = ''
    if data.name:
        log_input_string += f'Name: {document.name} => {data.name}; '
        document.name = data.name
    if data.expiry_date:
        log_input_string += f'Verfallsdatum: {document.expiry_date} => {data.expiry_date}; '
        document.expiry_date = data.expiry_date
    if data.delete_file:
        log_input_string += f'Dokument gelöscht: {document.file} '
        document.file.delete()
    if data.done:
        log_input_string += f'Erledigt ✅; '
        document.done = True
    if data.done is False:
        log_input_string += f'Nicht Erledigt ❌; '
        document.done = False
    document.save()
    log_input = f'Details vom LKW-Dokument/Termin {document.name} wurden verändert. Veränderungen: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=own_firm,
                       log_input=log_input)
    return 200


@api.post('/delete-truck-document', tags=['Truck'])
def delete_truck_document(request, document_id: int = Form(...)):
    document = get_object_or_404(TruckDocument, id=document_id)
    document.delete()
    return 200


@api.post('/add-truck-document-file', tags=['Truck'])
def add_truck_document_file(request, document_id: int, file: UploadedFile = File(...)):
    truck_document = TruckDocument.objects.get(id=document_id)
    truck_document.file = file
    truck_document.save()
    return 200


@api.post('/change-truck-payment', tags=['Truck'])
def change_truck_payment(request, data: ChangeTruckPayment = Form(...)):
    own_firm = get_object_or_404(OwnFirm, name=data.own_firm)
    truck = get_object_or_404(Truck, own_firm=own_firm,
                              plate=get_object_or_404(Truck, plate=data.plate, own_firm=own_firm))
    log_input_string = ''
    if data.payment_method:
        log_input_string += f'Zahlungsmethode: {if_null(truck.payment_method)} => {data.payment_method}; '
        truck.payment_method = data.payment_method
    if data.price:
        log_input_string += f'Preis: {if_null(truck.price)} => {data.price}; '
        truck.price = data.price
    if data.total_installment_months:
        log_input_string += f'Zeitraum: {if_null(truck.total_installment_months)} => {data.total_installment_months}; '
        truck.total_installment_months = data.total_installment_months
    if data.installment_monthly_payment_amount:
        log_input_string += f'Ratenbetrag: {if_null(truck.installment_monthly_payment_amount)} => {data.installment_monthly_payment_amount}; '
        truck.installment_monthly_payment_amount = data.installment_monthly_payment_amount
    if data.paid_status or data.paid_status == False:
        def verdeutschern(word):
            if word:
                return 'Bezahlt'
            else:
                return 'Zahlung Läuft'

        log_input_string += f'Zahlungsstatus: {verdeutschern(truck.paid_status)} => {verdeutschern(data.paid_status)}; '
        truck.paid_status = data.paid_status
    if data.paid_day:
        log_input_string += f'Bezahltes Datum: {if_null(datetime.strftime(truck.paid_day, "%d.%m.%Y"))} => {datetime.strftime(datetime.strptime(data.paid_day, "%Y-%m-%d"), "%d.%m.%Y")}; '
        truck.paid_day = data.paid_day
    if data.installment_start_date:
        log_input_string += f'Anfangsdatum: {if_null(datetime.strftime(truck.installment_start_date, "%d.%m.%Y"))} => {datetime.strftime(datetime.strptime(data.installment_start_date, "%Y-%m-%d"), "%d.%m.%Y")}; '
        truck.installment_start_date = data.installment_start_date
    if data.installment_end_date:
        log_input_string += f'Enddatum: {if_null(datetime.strftime(truck.installment_end_date, "%d.%m.%Y"))} => {datetime.strftime(datetime.strptime(data.installment_end_date, "%Y-%m-%d"), "%d.%m.%Y")}; '
        truck.installment_end_date = data.installment_end_date
    truck.save()
    log_input = f'Zahlungsdetails vom LKW {truck.plate} wurden verändert. Veränderungen: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=own_firm,
                       log_input=log_input)
    return 200


# CONTACT
@api.get('/get-contacts', tags=['Contact'], description='order_by options: name, firm, phone, mail')
def get_contacts(request, query: GetContactsSchema = Query(...)):
    if query.order_by:
        if query.order_by == 'name':
            if query.direction == 'asc':
                contacts = Contact.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'name')
            elif query.direction == 'des':
                contacts = Contact.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-name')
        elif query.order_by == 'firm':
            if query.direction == 'asc':
                contacts = Contact.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'firm')
            elif query.direction == 'des':
                contacts = Contact.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-firm')
        elif query.order_by == 'phone':
            if query.direction == 'asc':
                contacts = Contact.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'phone')
            elif query.direction == 'des':
                contacts = Contact.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-phone')
        elif query.order_by == 'mail':
            if query.direction == 'asc':
                contacts = Contact.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'mail')
            elif query.direction == 'des':
                contacts = Contact.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-mail')
    else:
        contacts = Contact.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
            '-id')
    if query.name:
        contacts = contacts.filter(name__contains=query.name)
    if query.firm:
        contacts = contacts.filter(firm__contains=query.firm)
    if query.tag:
        contacts = contacts.filter(tag=get_object_or_404(ContactTag, name=query.tag))

    contacts_return = []
    for contact in contacts:
        contact_details = {'id': contact.id, 'name': contact.name, 'firm': contact.firm, 'phone': contact.phone,
                           'mail': contact.mail, 'tag': str(contact.tag)}
        contacts_return.append(contact_details)

    return contacts_return


@api.post('/create-contact', tags=['Contact'])
def create_contact(request, data: CreateContactSchema = Form(...)):
    own_firm = get_object_or_404(OwnFirm, name=data.own_firm)
    created_contact = Contact.objects.create(own_firm=own_firm, name=data.name, firm=data.firm,
                                             phone=data.phone, mail=data.mail,
                                             tag=get_object_or_404(ContactTag, name=data.tag),
                                             address=data.address, note=data.note)

    log_input = f'Kontakt hinzugefügt: {created_contact.name}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=own_firm,
                       log_input=log_input)
    return 200


@api.get('/get-contact', tags=['Contact'], description='order_by options: date')
def get_contact(request, query: GetContact = Query(...)):
    contact = get_object_or_404(Contact, id=query.contact_id)
    contact_informations = {'id': contact.id, 'name': contact.name, 'firm': contact.firm, 'phone': contact.phone,
                            'mail': contact.mail, 'tag': str(contact.tag), 'fax': contact.fax,
                            'address': contact.address, 'notes': contact.address}
    meetings = []
    if query.order_by:
        if query.order_by == 'date':
            if query.direction == 'asc':
                qs = Meeting.objects.filter(contact=contact).order_by('meetings_date')
            elif query.direction == 'des':
                qs = Meeting.objects.filter(contact=contact).order_by('-meetings_date')
    else:
        qs = Meeting.objects.filter(contact=contact).order_by('-id')
    for meeting in qs:
        meeting_details = {'id': meeting.id, 'meetings_date': meeting.meetings_date,
                           'meetings_notes': meeting.meetings_notes}
        meetings.append(meeting_details)

    return {'details': contact_informations, 'meetings': meetings}


@api.post('/update-contact', tags=['Contact'])
def update_contact(request, data: UpdateContactSchema = Form(...)):
    contact = get_object_or_404(Contact, id=data.contact_id)
    log_input_string = ''
    if data.name:
        log_input_string += f'Name: {contact.name} => {data.name}; '
        contact.name = data.name
    if data.firm:
        log_input_string += f'Firma: {if_null(contact.firm)} => {data.firm}; '
        contact.firm = data.firm
    if data.tag:
        tag = get_object_or_404(ContactTag, name=data.tag)
        log_input_string += f'Etikett: {if_null(contact.tag)} => {tag.name}; '
        contact.tag = tag
    if data.phone:
        log_input_string += f'Telefonnummer: {if_null(contact.phone)} => {data.phone}; '
        contact.phone = data.phone
    if data.mail:
        log_input_string += f'Mail: {if_null(contact.mail)} => {data.mail}; '
        contact.mail = data.mail
    if data.fax:
        log_input_string += f'Fax: {if_null(contact.fax)} => {data.fax}; '
        contact.fax = data.fax
    if data.address:
        log_input_string += f'Adresse: {if_null(contact.address)} => {data.address}; '
        contact.address = data.address
    if data.note:
        log_input_string += f'Notizen: {if_null(contact.note)} => {data.note}; '
        contact.note = data.note
    contact.save()

    log_input = f'Daten zum Kontakt {contact.name} wurden verändert. Veränderungen: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=contact.own_firm,
                       log_input=log_input)
    return 200


@api.post('/delete-contact', tags=['Contact'])
def delete_contact(request, contact_id: int = Form(...), admin: str = Form(...)):
    contact = get_object_or_404(Contact, id=contact_id)
    contact.delete()
    log_input = f'Kontakt gelöscht: {contact.name}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=contact.own_firm,
                       log_input=log_input)
    return 200


@api.post('/create-meeting', tags=['Contact'])
def create_meeting(request, data: CreateMeetingSchema = Form(...)):
    contact = get_object_or_404(Contact, id=data.contact)
    date = datetime.strptime(data.meetings_date, "%Y-%m-%d").date()
    try:
        created_meeting = Meeting.objects.create(meetings_notes=data.meetings_notes, meetings_date=date,
                                                 contact=contact)
    except IntegrityError:
        return HttpResponse('Meeting has to be associated with Contact', status=406)
    log_input = f'Meeting am {created_meeting.meetings_date.strftime("%d.%m.%Y")} mit dem Kontakt {contact.name} wurde hinzugefügt'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=contact.own_firm,
                       log_input=log_input)
    return 200


@api.post('/delete-meeting', tags=['Contact'])
def delete_meeting(request, meeting_id: int = Form(...), admin: str = Form(...)):
    meeting = get_object_or_404(Meeting, id=meeting_id)
    meeting.delete()
    log_input = f'Meeting am {meeting.meetings_date.strftime("%d.%m.%Y")} mit {meeting.contact.name} wurde gelöscht.'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=meeting.contact.own_firm,
                       log_input=log_input)
    return 200


@api.post('/update-meeting', tags=['Contact'])
def update_meeting(request, data: UpdateMeeting = Form(...)):
    meeting = get_object_or_404(Meeting, id=data.meeting_id)
    log_input_string = ''
    if data.meetings_notes:
        log_input_string += f'Notizen: {meeting.meetings_notes} => {data.meetings_notes}; '
        meeting.meetings_notes = data.meetings_notes
    if data.meetings_date:
        date = datetime.strptime(data.meetings_date, "%Y-%m-%d").date()
        log_input_string += f'Datum: {meeting.meetings_date.strftime("%d.%m.%Y")} => {date.strftime("%d.%m.%Y")}; '
        meeting.meetings_date = date
    meeting.save()
    log_input = f'Daten zum Meeting mit {meeting.contact.name} wurden verändert. Veränderungen: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=meeting.contact.own_firm,
                       log_input=log_input)
    return 200


# FUELCARD
@api.get('/get-fuelcards', tags=['Fuelcard'], description='order_by options: card_nr, firm, driver')
def get_fuelcards(request, query: GetFuelCardsSchema = Query(...)):
    if query.order_by:
        if query.order_by == 'card_nr':
            if query.direction == 'asc':
                qs = Fuelcard.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'card_nr')
            elif query.direction == 'des':
                qs = Fuelcard.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-card_nr')
        elif query.order_by == 'firm':
            if query.direction == 'asc':
                qs = Fuelcard.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'firm__name')
            elif query.direction == 'des':
                qs = Fuelcard.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-firm__name')
        elif query.order_by == 'driver':
            if query.direction == 'asc':
                qs = Fuelcard.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'card_is_at__name')
            elif query.direction == 'des':
                qs = Fuelcard.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-card_is_at__name')
    else:
        qs = Fuelcard.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-id')
    if query.firm:
        qs = qs.filter(firm=get_object_or_404(FuelcardFirm, id=query.firm))
    if query.driver:
        qs = qs.filter(card_is_at=get_object_or_404(Worker, id=query.driver))
    if query.active:
        qs = qs.filter(status=True)
    fuelcards = []
    for fuelcard in qs:

        def get_firm():
            if fuelcard.firm:
                return fuelcard.firm.name
            else:
                return None

        fuelcard_details = {'id': fuelcard.id, 'card_nr': fuelcard.card_nr, 'firm': get_firm(),
                            'card_is_at': str(fuelcard.card_is_at)}
        if fuelcard.status:
            fuelcard_details['status'] = 'Aktiv'
        else:
            fuelcard_details['status'] = 'Inaktiv'
        fuelcards.append(fuelcard_details)
    return fuelcards


@api.post('/create-fuelcard', tags=['Fuelcard'])
def create_fuelcard(request, data: CreateFuelCardSchema = Form(...)):
    def status_func():
        if data.status == 'Aktiv':
            return True
        elif data.status == 'Inaktiv':
            return False

    fuelcard_firm, created = FuelcardFirm.objects.get_or_create(name=data.firm)

    try:
        created_fuelcard = Fuelcard.objects.create(own_firm=get_object_or_404(OwnFirm, name=data.own_firm),
                                                   firm=fuelcard_firm,
                                                   card_nr=data.card_nr,
                                                   notes=data.notes, status=status_func())
    except IntegrityError:
        return HttpResponse('Es existiert bereits eine Tankkarte mit dieser Nummer', status=406)
    log_input = f'Tankkarte {created_fuelcard.card_nr} wurde hinzugefügt'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=created_fuelcard.own_firm,
                       log_input=log_input)
    return 200


@api.get('/get-fuelcard', tags=['Fuelcard'], description='order_by options: driver, got_date, gave_back_date')
def get_fuelcard(request, query: GetFuelcard = Query(...)):
    fuelcard = get_object_or_404(Fuelcard, id=query.card_id)
    if fuelcard.firm:
        fuelcard_firm_name = fuelcard.firm.name
        fuelcard_firm_id = fuelcard.firm.id
    else:
        fuelcard_firm_name = None
        fuelcard_firm_id = None
    fuelcard_details = {'card_nr': fuelcard.card_nr, 'firm_id': fuelcard_firm_id,
                        'firm_name': fuelcard_firm_name}
    if fuelcard.status:
        fuelcard_details['status'] = 'Aktiv'
    else:
        fuelcard_details['status'] = 'Inaktiv'
    fuelcard_details['notes'] = 'Inaktiv'
    fuelcard_details['notes'] = fuelcard.notes
    fuelcard_activities = []
    if query.order_by:
        if query.order_by == 'driver':
            if query.direction == 'asc':
                qs = FuelcardActivities.objects.filter(fuelcard=fuelcard).order_by('driver__name')
            elif query.direction == 'des':
                qs = FuelcardActivities.objects.filter(fuelcard=fuelcard).order_by('-driver__name')
        elif query.order_by == 'got_date':
            if query.direction == 'asc':
                qs = FuelcardActivities.objects.filter(fuelcard=fuelcard).order_by('got_date')
            elif query.direction == 'des':
                qs = FuelcardActivities.objects.filter(fuelcard=fuelcard).order_by('-got_date')
        elif query.order_by == 'gave_back_date':
            if query.direction == 'asc':
                qs = FuelcardActivities.objects.filter(fuelcard=fuelcard).order_by('gave_back_date')
            elif query.direction == 'des':
                qs = FuelcardActivities.objects.filter(fuelcard=fuelcard).order_by('-gave_back_date')
    else:
        qs = FuelcardActivities.objects.filter(fuelcard=fuelcard).order_by('-id')
    for activity in qs:
        activity_details = {'id': activity.id, 'driver': str(activity.driver), 'got_date': activity.got_date,
                            'gave_back_date': activity.gave_back_date}
        fuelcard_activities.append(activity_details)
    return {'details': fuelcard_details, 'activities': fuelcard_activities}


@api.post('/delete-fuelcard', tags=['Fuelcard'])
def delete_fuelcard(request, card_id: int = Form(...), admin: str = Form(...)):
    fuelcard = get_object_or_404(Fuelcard, id=card_id)
    fuelcard.delete()
    log_input = f'Tankkarte {fuelcard.card_nr} wurde gelöscht'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=fuelcard.own_firm,
                       log_input=log_input)
    return 200


@api.post('/update-fuelcard', tags=['Fuelcard'])
def update_fuelcard(request, data: UpdateFuelcardSchema = Form(...)):
    fuelcard = get_object_or_404(Fuelcard, id=data.card_id)
    log_input_string = ''
    if data.card_nr:
        if Fuelcard.objects.filter(card_nr=data.card_nr).exists():
            return HttpResponse('Es existiert bereits eine Tankkarte mit dieser Nummer', status=406)
        log_input_string += f'Kartennummer: {fuelcard.card_nr} => {data.card_nr} ; '
        fuelcard.card_nr = data.card_nr
    if data.firm:
        firm, created = FuelcardFirm.objects.get_or_create(name=data.firm)
        log_input_string += f'Firmenname: {fuelcard.firm} => {firm.name} ; '
        fuelcard.firm = firm
    if data.status:
        if data.status == 'Aktiv':
            log_input_string += f'Status: {fuelcard.status} => Aktiv ; '
            fuelcard.status = True
        elif data.status == 'Inaktiv':
            log_input_string += f'Status: {fuelcard.status} => Inaktiv ; '
            fuelcard.status = False
    if data.notes:
        log_input_string += f'Notizen: {fuelcard.notes} => {data.notes} ; '
        fuelcard.notes = data.notes

    fuelcard.save()
    log_input = f'Daten zur Tankkarte {fuelcard.card_nr} wurden verändert. Veränderungen: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=fuelcard.own_firm,
                       log_input=log_input)
    return 200


@api.post('/change-fuelcard-driver', tags=['Fuelcard'])
def change_fuelcard_driver(request, data: ChangeFuelCardDriverSchema = Form(...)):
    fuelcard = get_object_or_404(Fuelcard, id=data.card_id)
    driver = get_object_or_404(Worker, id=data.worker_id)
    a = FuelcardActivities.objects.filter(fuelcard=fuelcard)
    if a:
        latest_a = a.order_by('got_date').last()
        latest_a.gave_back_date = data.got_date
        latest_a.save()
    created_activity = FuelcardActivities.objects.create(fuelcard=fuelcard, driver=driver, got_date=data.got_date,
                                                         gave_back_date=data.gave_back_date)
    fuelcard.card_is_at = driver
    fuelcard.save()
    log_input = f'Tankkarte {fuelcard.card_nr} wurde am {datetime.strftime(datetime.strptime(created_activity.got_date, "%Y-%m-%d"), "%d.%m.%Y")} an {driver.name} abgegeben'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=fuelcard.own_firm,
                       log_input=log_input)
    return 200


# @api.post('/update-fuelcard-activity', tags=['Fuelcard'])
# def update_fuelcard_activity(request, data: UpdateFuelcardActivity = Form(...)):
#     pass


@api.post('/delete-fuelcard-activity', tags=['Fuelcard'])
def delete_fuelcard_activity(request, activity_id: int = Form(...)):
    activity = get_object_or_404(FuelcardActivities, id=activity_id)
    activity.delete()
    return 200


# BILL
@api.get('/get-bills', tags=['Bill'],
         description='order_by options: bill_nr, firm, creation_date, has_to_be_paid_date, sum, end_sum')
def get_bills(request, query: GetBillsSchema = Query(...)):
    if query.order_by:
        if query.order_by == 'bill_nr':
            if query.direction == 'asc':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('bill_nr')
            elif query.direction == 'des':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-bill_nr')
        elif query.order_by == 'firm':
            if query.direction == 'asc':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'firm__name')
            elif query.direction == 'des':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-firm__name')
        elif query.order_by == 'creation_date':
            if query.direction == 'asc':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'creation_date')
            elif query.direction == 'des':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-creation_date')
        elif query.order_by == 'has_to_be_paid_date':
            if query.direction == 'asc':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'has_to_be_paid_date_start')
            elif query.direction == 'des':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-has_to_be_paid_date_start')
        elif query.order_by == 'sum':
            if query.direction == 'asc':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('sum')
            elif query.direction == 'des':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-sum')
        elif query.order_by == 'end_sum':
            if query.direction == 'asc':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('end_sum')
            elif query.direction == 'des':
                qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-end_sum')
    else:
        qs = Bill.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-id')
    if query.firm:
        qs = qs.filter(firm=get_object_or_404(Firm, name=query.firm))
    if query.netto_start:
        qs = qs.filter(end_sum__gte=query.netto_start)
    if query.netto_end:
        qs = qs.filter(end_sum__lte=query.netto_end)
    if query.brutto_start:
        qs = qs.filter(sum__gte=query.brutto_start)
    if query.brutto_end:
        qs = qs.filter(sum__lte=query.brutto_end)
    if query.created_date_start:
        qs = qs.filter(creation_date__gte=query.created_date_start)
    if query.created_date_end:
        qs = qs.filter(creation_date__lte=query.created_date_end)
    bills_list = []
    for bill in qs:
        def firm_name():
            if bill.firm:
                return bill.firm.name
            else:
                return None

        bill_details = {'id': bill.id, 'bill_nr': bill.bill_nr, 'firm': firm_name(),
                        'creation_date': bill.creation_date,
                        'has_to_be_paid_date_start': bill.has_to_be_paid_date_start,
                        'has_to_be_paid_date_end': bill.has_to_be_paid_date_end,
                        'end_sum': bill.end_sum, 'sum': bill.sum}
        if bill.pdf:
            bill_details['file'] = root + bill.pdf.url
        bills_list.append(bill_details)
    return bills_list


@api.post('/delete-bill', tags=['Bill'])
def delete_bill(request, bill_id: int = Form(...), admin: str = Form(...)):
    bill = get_object_or_404(Bill, id=bill_id)
    bill.delete()
    log_input = f'Rechnung {bill.bill_nr} wurde gelöscht'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=bill.own_firm,
                       log_input=log_input)
    return 200


@api.post('/create-bill', tags=['Bill'])
def create_bill(request, data: CreateBillSchema = Form(...)):
    own_firm = get_object_or_404(OwnFirm, name=data.own_firm)

    creation_date = datetime.strptime(data.creation_date, '%Y-%m-%d')

    def bill_nr_generator():
        bills = Bill.objects.filter(own_firm=own_firm, creation_date__year=creation_date.year)
        if bills.exists():
            highest_nr_bill = bills.order_by('-bill_nr').first()
            if highest_nr_bill.bill_nr:
                bill_nr_int = highest_nr_bill.bill_nr_int + 1
                bill_nr = f'{bill_nr_int:03d}/{creation_date.year}'
            else:
                bill_nr_int = 1
                bill_nr = f'001/{creation_date.year}'
        else:
            bill_nr_int = 1
            bill_nr = f'001/{creation_date.year}'
        return bill_nr_int, bill_nr

    bill_nr = bill_nr_generator()

    firm_obj, created = Firm.objects.get_or_create(own_firm=own_firm, name=data.firm)
    created_bill = Bill.objects.create(own_firm=own_firm,
                                       firm=firm_obj, bill_nr_int=bill_nr[0],
                                       bill_nr=bill_nr[1],
                                       customer_tax_nr=data.customer_tax_nr, address=data.address,
                                       creation_date=creation_date,
                                       has_to_be_paid_date_start=data.has_to_be_paid_date_start,
                                       has_to_be_paid_date_end=data.has_to_be_paid_date_end,
                                       taxes=data.taxes
                                       )
    products = json.loads(data.products)
    for product in products:
        product_sum = Decimal(product['amount'] * product['unit_price'])
        created_bill.products.add(
            Product.objects.create(position=product['position'], description=product['description'],
                                   amount=product['amount'], unit=product['unit'], unit_price=product['unit_price'],
                                   sum=product_sum))
        if created_bill.sum:
            created_bill.sum += product_sum
        else:
            created_bill.sum = product_sum
    created_bill.end_sum = created_bill.sum * ((Decimal(100) + Decimal(created_bill.taxes)) / Decimal(100))
    created_bill.save()
    log_input = f'Rechnung {created_bill.bill_nr} wurde erstellt'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=created_bill.own_firm,
                       log_input=log_input)

    paytime = False
    pay_start = None
    pay_end = None
    if created_bill.has_to_be_paid_date_start and created_bill.has_to_be_paid_date_end:
        paytime = True
        if datetime.strptime(created_bill.has_to_be_paid_date_start, '%Y-%m-%d').year == datetime.strptime(
                created_bill.has_to_be_paid_date_end, '%Y-%m-%d').year:
            pay_start = datetime.strftime(datetime.strptime(created_bill.has_to_be_paid_date_start, '%Y-%m-%d'),
                                          '%d.%m')
        else:
            pay_start = datetime.strftime(datetime.strptime(created_bill.has_to_be_paid_date_start, '%Y-%m-%d'),
                                          '%d.%m.%Y')
        pay_end = datetime.strftime(datetime.strptime(created_bill.has_to_be_paid_date_end, '%Y-%m-%d'), '%d.%m.%Y')

    produkte = []
    for product in products:
        next_line = False
        if len(product['description']) > 42:
            next_line = True
        produkte.append({'pos': product['position'], 'beschreibung': product['description'],
                         'next_line': next_line,
                         'menge': format_number(float(product['amount'])), 'einheit': product['unit'],
                         'e_preis': format_number(float(product['unit_price'])),
                         'summe': format_number((float(product['unit_price']) * float(product['amount'])))})

    context = {
        'own_firm': own_firm,
        'customer': firm_obj,
        'bill_nr': created_bill.bill_nr,
        'bill_date': datetime.strftime(created_bill.creation_date, '%d.%m.%Y'),
        'paytime': paytime,
        'pay_start': pay_start,
        'pay_end': pay_end,
        'produkte': produkte,
        'zwischensumme': format_number(created_bill.sum),
        'mwtprozent': f'{created_bill.taxes:g}',
        'mwtsumme': format_number(float(created_bill.sum) * (created_bill.taxes / 100)),
        'endsumme': format_number(created_bill.end_sum)
    }

    html = render_to_string('rechnung.html', context)
    pdf_file = HTML(string=html).write_pdf()
    file_name = f"Rechnung -- {created_bill.own_firm} -- {created_bill.bill_nr}.pdf"
    created_bill.pdf.save(file_name, ContentFile(pdf_file))
    created_bill.save()
    base64_pdf = base64.b64encode(pdf_file).decode('utf-8')
    response = {'name': file_name, 'file': base64_pdf}
    return response


@api.post('/add-bill-document', tags=['Bill'])
def add_bill_document(request, bill_id: str = Form(...), file: UploadedFile = File(...)):
    bill = get_object_or_404(Bill, bill_nr=bill_id)
    bill.pdf = file
    bill.save()
    return 200


# GUTSCHRIFT
@api.get('/get-gutschrifts', tags=['Gutschrift'], description='order_by options: avis, document_nr, firm, open_amount')
def get_gutschrifts(request, query: GetGutschrifts = Query(...)):
    if query.order_by:
        if query.order_by == 'avis':
            if query.direction == 'asc':
                qs = Gutschrift.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'avis')
            elif query.direction == 'des':
                qs = Gutschrift.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-avis')
        elif query.order_by == 'document_nr':
            if query.direction == 'asc':
                qs = Gutschrift.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'document_nr')
            elif query.direction == 'des':
                qs = Gutschrift.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-document_nr')
        elif query.order_by == 'firm':
            if query.direction == 'asc':
                qs = Gutschrift.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'firm__name')
            elif query.direction == 'des':
                qs = Gutschrift.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-firm__name')
        elif query.order_by == 'open_amount':
            if query.direction == 'asc':
                qs = Gutschrift.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'open_amount')
            elif query.direction == 'des':
                qs = Gutschrift.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-open_amount')
    else:
        qs = Gutschrift.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-id')
    if query.firm:
        firm = get_object_or_404(Firm, id=query.firm)
        qs = qs.filter(firm=firm)
    if query.avis_nr:
        qs = qs.filter(avis__icontains=query.avis_nr)
    if query.document_nr:
        qs = qs.filter(document_nr__icontains=query.document_nr)
    if query.paid_start:
        qs = qs.filter(paid_amount__gte=query.paid_start)
    if query.paid_end:
        qs = qs.filter(paid_amount__lte=query.paid_end)
    if query.open_start:
        qs = qs.filter(open_amount__gte=query.open_start)
    if query.open_end:
        qs = qs.filter(open_amount__lte=query.open_end)
    gutschriften = []

    def get_firm_name():
        if gutschrift.firm:
            return gutschrift.firm.name
        else:
            return None

    def get_file_url():
        if gutschrift.file:
            return root + gutschrift.file.url
        else:
            return None

    for gutschrift in qs:
        details = {'id': gutschrift.id, 'avis_nr': gutschrift.avis, 'document_nr': gutschrift.document_nr,
                   'firm_name': get_firm_name(), 'paid_amount': gutschrift.paid_amount,
                   'open_amount': gutschrift.open_amount, 'file_url': get_file_url()}
        gutschriften.append(details)

    return gutschriften


@api.post('/create-gutschrift', tags=['Gutschrift'])
def create_gutschrift(request, data: CreateGutschrift = Form(...)):
    own_firm = get_object_or_404(OwnFirm, name=data.own_firm)
    firm = get_object_or_404(Firm, id=data.firm)

    def get_open_amount():
        return data.gross_amount * ((100 + data.taxes) / 100)

    try:
        created_gutschrift = Gutschrift.objects.create(own_firm=own_firm, firm=firm, creation_date=data.creation_date,
                                                       start=data.start, end=data.end,
                                                       document_nr=data.document_nr, gross_amount=data.gross_amount,
                                                       taxes=data.taxes, open_amount=get_open_amount(),
                                                       avis=data.avis_nr)
    except IntegrityError:
        return HttpResponse('Es existiert bereits eine Gutschrift mit dieser Dokument Nr.', status=406)
    log_input = f'Gutschrift {created_gutschrift.document_nr} wurde erstellt.'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=own_firm,
                       log_input=log_input)
    return created_gutschrift.id


@api.post('/add-gutschrift-document', tags=['Gutschrift'])
def add_gutschrift_document(request, gutschrift_id: int = Form(...), file: UploadedFile = File(...)):
    gutschrift = Gutschrift.objects.get(id=gutschrift_id)
    gutschrift.file = file
    gutschrift.save()
    return 200


@api.get('/get-gutschrift', tags=['Gutschrift'], description='order_by options: amount, date')
def get_gutschrift(request, query: GetGutschrift = Query(...)):
    gutschrift = get_object_or_404(Gutschrift, id=query.gutschrift_id)

    def get_firm():
        if gutschrift.firm:
            return gutschrift.firm.name
        else:
            return None

    def get_file_url():
        if gutschrift.file:
            return root + gutschrift.file.url
        else:
            return None

    def get_taxes():
        if gutschrift.taxes:
            return Decimal(gutschrift.taxes)
        else:
            return None

    gutschrift_payments = []

    if GutschriftPayment.objects.filter(gutschrift=gutschrift).exists():
        if query.order_by:
            if query.order_by == 'amount':
                if query.direction == 'asc':
                    qs = GutschriftPayment.objects.filter(gutschrift=gutschrift).order_by('amount')
                elif query.direction == 'des':
                    qs = GutschriftPayment.objects.filter(gutschrift=gutschrift).order_by('-amount')
            elif query.order_by == 'date':
                if query.direction == 'asc':
                    qs = GutschriftPayment.objects.filter(gutschrift=gutschrift).order_by(
                        'date')
                elif query.direction == 'des':
                    qs = GutschriftPayment.objects.filter(gutschrift=gutschrift).order_by(
                        '-date')
        else:
            qs = GutschriftPayment.objects.filter(gutschrift=gutschrift).order_by('-id')
        for gp in qs:
            gp_details = {'id': gp.id, 'amount': gp.amount, 'date': gp.date}
            gutschrift_payments.append(gp_details)

    gutschrift_details = {'id': gutschrift.id, 'document_nr': gutschrift.document_nr, 'avis': gutschrift.avis,
                          'firm': get_firm(),
                          'creation_date': str(gutschrift.creation_date), 'start': gutschrift.start,
                          'end': gutschrift.end, 'taxes': get_taxes(), 'paid': gutschrift.paid_amount,
                          'open': gutschrift.open_amount,
                          'file_url': get_file_url(), 'gutschrift_payments': gutschrift_payments}

    return gutschrift_details


@api.post('/update-gutschrift', tags=['Gutschrift'])
def update_gutschrift(request, data: UpdateGutschrift = Form(...)):
    gutschrift = get_object_or_404(Gutschrift, id=data.gutschrift_id)
    log_input_string = ''
    if data.document_nr:
        if Gutschrift.objects.filter(document_nr=data.document_nr).exists():
            return HttpResponse('Es existiert bereits eine Gutschrift mit dieser Dokument Nr.', status=406)
        log_input_string += f'Dokumentnummer: {gutschrift.document_nr} => {data.document_nr}'
        gutschrift.document_nr = data.document_nr
    if data.firm:
        firm = get_object_or_404(Firm, id=data.firm)
        log_input_string += f'Firmenname: {if_null(gutschrift.firm.name)} => {firm.name}'
        gutschrift.firm = firm
    if data.creation_date:
        log_input_string += f'Eingangsdatum: {if_null(gutschrift.creation_date.strftime("%d.%m.%Y"))} => {datetime.strftime(datetime.strptime(data.creation_date, "%Y-%m-%d"), "%d.%m.%Y")}'
        gutschrift.creation_date = data.creation_date
    if data.start:
        log_input_string += f'Zeitraum-Start: {if_null(gutschrift.start.strftime("%d.%m.%Y"))} => {datetime.strftime(datetime.strptime(data.start, "%Y-%m-%d"), "%d.%m.%Y")}'
        gutschrift.start = data.start
    if data.end:
        log_input_string += f'Zeitraum-Ende: {if_null(gutschrift.end.strftime("%d.%m.%Y"))} => {datetime.strftime(datetime.strptime(data.end, "%Y-%m-%d"), "%d.%m.%Y")}'
        gutschrift.end = data.end
    if data.taxes:
        log_input_string += f'MwSt. : %{if_null(gutschrift.taxes)} => %{data.taxes}'
        gutschrift.taxes = Decimal(data.taxes)
    if data.gross_amount:
        gutschrift.gross_amount = Decimal(data.gross_amount)
        gutschrift.open_amount = gutschrift.gross_amount * Decimal(((100 + gutschrift.taxes) / 100)) - Decimal(
            gutschrift.paid_amount)
        log_input_string += f'Brutto. : {if_null(gutschrift.gross_amount)} => {data.gross_amount}'
    if data.avis_nr:
        log_input_string += f'Avis Nr. : {if_null(gutschrift.avis)} => {data.avis_nr}'
        gutschrift.avis = data.avis_nr
    gutschrift.save()
    log_input = f''
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=gutschrift.own_firm,
                       log_input=log_input)
    return 200


@api.post('/update-gutschrift-file', tags=['Gutschrift'])
def update_gutschrift_file(request, gutschrift_id: int = Form(...), admin: str = Form(...),
                           file: UploadedFile = File(...)):
    gutschrift = get_object_or_404(Gutschrift, id=gutschrift_id)
    gutschrift.file = file
    gutschrift.save()
    log_input = f'Dokument von der Gutschrift {gutschrift.document_nr} wurde verändert'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=gutschrift.own_firm,
                       log_input=log_input)
    return 200


@api.post('/delete-gutschrift', tags=['Gutschrift'])
def delete_gutschrift(request, gutschrift_id: int = Form(...), admin: str = Form(...)):
    gutschrift = get_object_or_404(Gutschrift, id=gutschrift_id)
    gutschrift.delete()
    log_input = f'Gutschrift {gutschrift.document_nr} wurde gelöscht'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=gutschrift.own_firm,
                       log_input=log_input)
    return 200


@api.post('/add-gutschrift-payment', tags=['Gutschrift'])
def add_gutschrift_payment(request, data: CreateGutschriftPayment = Form(...)):
    gutschrift = get_object_or_404(Gutschrift, id=data.gutschrift_id)
    GutschriftPayment.objects.create(gutschrift=gutschrift, amount=Decimal(data.amount), date=data.date)
    gutschrift.open_amount -= Decimal(data.amount)
    gutschrift.paid_amount += Decimal(data.amount)
    if gutschrift.open_amount <= 0:
        gutschrift.completely_paid = True
        gutschrift.completely_paid_date = today
    gutschrift.save()
    log_input = f'Neue Zahlung im Wert von {data.amount} € wurde zur Gutschrift {gutschrift.document_nr} hinzugefügt'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=gutschrift.own_firm,
                       log_input=log_input)
    return 200


@api.post('/delete-gutschrift-payment', tags=['Gutschrift'])
def delete_gutschrift_payment(request, gutschrift_payment_id: int = Form(...), admin: str = Form(...)):
    gutschrift = get_object_or_404(GutschriftPayment, id=gutschrift_payment_id)
    gutschrift.delete()
    log_input = f'Gutschrift Zahlung für {gutschrift.gutschrift.document_nr} wurde gelöscht'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=gutschrift.gutschrift.own_firm,
                       log_input=log_input)
    return 200


# WORKER
@api.get('/get-workers', tags=['Worker'],
         description='order_by options: name, holidays, remaining_holidays, salary, daily_expense')
def get_workers(request, query: GetWorkers = Query(...)):
    own_firm = get_object_or_404(OwnFirm, name=query.own_firm)
    worker_list = []
    if query.show_quit_workers:
        qs = Worker.objects.filter(own_firm=own_firm)
    else:
        qs = Worker.objects.filter(own_firm=own_firm, is_working=True)
    if query.order_by:
        if query.order_by == 'name':
            if query.direction == 'asc':
                qs = qs.order_by('name')
            elif query.direction == 'des':
                qs = qs.order_by('-name')
        elif query.order_by == 'holidays':
            if query.direction == 'asc':
                qs = qs.order_by('holidays')
            elif query.direction == 'des':
                qs = qs.order_by('-holidays')
        elif query.order_by == 'remaining_holidays':
            if query.direction == 'asc':
                qs = qs
            elif query.direction == 'des':
                qs = qs
        elif query.order_by == 'salary':
            if query.direction == 'asc':
                qs = qs.order_by('salary')
            elif query.direction == 'des':
                qs = qs.order_by('-salary')
        elif query.order_by == 'daily_expense':
            if query.direction == 'asc':
                qs = qs.order_by('daily_expense')
            elif query.direction == 'des':
                qs = qs.order_by('-daily_expense')
    for worker in qs:
        if HolidayAccount.objects.filter(worker=worker, year=today.year).exists():
            obj = HolidayAccount.objects.filter(worker=worker, year=today.year)[0]
            remaining_holidays = obj.remaining_holiday_days
        else:
            remaining_holidays = worker.holidays
        worker_details = {'id': worker.id, 'name': worker.name, 'holidays': worker.holidays,
                          'remaining_holidays': remaining_holidays, 'salary': worker.salary,
                          'daily_expense': worker.daily_expense}
        worker_list.append(worker_details)
    if query.order_by:
        if query.order_by == 'remaining_holidays':
            if query.direction == 'asc':
                worker_list = sorted(worker_list, key=lambda d: d['remaining_holidays'])
            elif query.direction == 'des':
                worker_list = sorted(worker_list, key=lambda d: d['remaining_holidays'], reverse=True)
    return worker_list


@api.get('/get-workers-list-excel', tags=['Worker'])
def get_workers_list_excel(request, own_firm: str = Query(...)):
    own_firm = get_object_or_404(OwnFirm, name=own_firm)
    data = {
        'index': [],
        'Personalnr.': [],
        'Name': [],
        'Eintritt': [],
        'Brutto': [],
        'Urlaubsanspruch': [],
        'Rest': [],
        'Krank': [],
        'Spesen': [],
        'Bemerkung': []
    }
    n = 1

    def if_not(elm):
        if not elm:
            return ''
        else:
            return elm

    for worker in Worker.objects.filter(own_firm=own_firm):
        data['index'].append(n)
        data['Personalnr.'].append(if_not(worker.worker_id))
        data['Name'].append(if_not(worker.name))
        data['Eintritt'].append(if_not(worker.start_date.strftime('%d.%m.%Y')))
        data['Brutto'].append(if_not(worker.salary))
        data['Urlaubsanspruch'].append(if_not(worker.holidays))
        if HolidayAccount.objects.filter(worker=worker, year=today.year).exists():
            holiday_account = HolidayAccount.objects.get(worker=worker, year=today.year)
            remaining = holiday_account.remaining_holiday_days
        else:
            remaining = worker.holidays
        data['Rest'].append(if_not(remaining))
        k = 0
        for od in Offday.objects.filter(worker=worker, tag__name='Urlaub', date__year=today.year):
            k += 1
        data['Krank'].append(if_not(k))
        data['Spesen'].append(if_not(worker.daily_expense))
        data['Bemerkung'].append('')
        n += 1

    df = pd.DataFrame(data)

    def get_total(i):
        k = 0
        for n in data[i]:
            if n:
                k += n
        return k

    df.loc['Blank'] = ''
    df.loc['Total'] = ['Total', '', '', '', get_total('Brutto'), '', '', get_total('Krank'), get_total('Spesen'), '']
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['J'].width = 30
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename="{own_firm.name} Arbeiterliste {today.month:02d}/{today.year}.xlsx"'
    file_name = f"{own_firm.name} Arbeiterliste {today.month:02d}/{today.year}.xlsx"
    workbook.save(response)
    response_content = response.content
    base64_content = base64.b64encode(response_content).decode('utf-8')
    response = {'name': file_name, 'file': base64_content}
    return response


@api.post('/create-worker', tags=['Worker'])
def create_worker(request, data: CreateWorker = Form(...)):
    own_firm = get_object_or_404(OwnFirm, name=data.own_firm)

    def get_position():
        try:
            position = get_object_or_404(Position, name=data.position)
        except Http404:
            position = Position.objects.create(name=data.position)
        return position

    def get_daily_expense():
        if data.daily_expenses:
            return Decimal(data.daily_expenses)
        else:
            return 0

    def get_salary():
        if data.salary:
            return Decimal(data.salary)
        else:
            return 0

    try:
        worker = Worker.objects.create(own_firm=own_firm, name=data.name, worker_id=data.worker_id,
                                       position=get_position(),
                                       start_date=data.start_date, salary=get_salary(),
                                       daily_expense=get_daily_expense(), holidays=data.holidays,
                                       note=data.note)
    except IntegrityError:
        return HttpResponse('Es existiert bereits ein Mitarbeiter mit dieser Personalnummer', status=406)

    if worker.position.name == 'Fahrer':
        worker.is_driver = True
    else:
        worker.is_driver = False
    if data.quit_date:
        worker.has_quit = True
        worker.quit_date = data.quit_date
    if data.remaining_holidays:
        HolidayAccount.objects.create(worker=worker, remaining_holiday_days=data.remaining_holidays, year=today.year)
    worker.save()
    log_input = f'Neuer Mitarbeiter {worker.name}({worker.worker_id}) wurde hinzugefügt'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=own_firm,
                       log_input=log_input)
    return 200


@api.get('/get-worker', tags=['Worker'])
def get_worker(request, worker_id: int):
    worker = get_object_or_404(Worker, id=worker_id)

    def get_remaining_holidays():
        if HolidayAccount.objects.filter(worker=worker, year=today.year).exists():
            ha = HolidayAccount.objects.filter(worker=worker, year=today.year)[0]
            return ha.remaining_holiday_days
        else:
            return None

    worker_details = {'id': worker.id, 'name': worker.name, 'worker_id': worker.worker_id,
                      'position': worker.position.name, 'salary': worker.salary, 'daily_expense': worker.daily_expense,
                      'holidays': worker.holidays, 'remaining_holidays': get_remaining_holidays(),
                      'start_date': worker.start_date, 'quit_date': worker.quit_date, 'note': worker.note}

    if DebtPayment.objects.filter(worker=worker).exists():
        debt_payments = []
        for debt in DebtPayment.objects.filter(worker=worker):
            debt_details = {'id': debt.id, 'amount': debt.amount, 'date': debt.date, 'notes': debt.notes}
            debt_payments.append(debt_details)
        worker_details['debt_payments'] = debt_payments

    if WorkerDocument.objects.filter(worker=worker).exists():
        worker_documents = []
        for worker_document in WorkerDocument.objects.filter(worker=worker):
            def get_file_url():
                if worker_document.file:
                    return root + worker_document.file.url
                else:
                    return None

            worker_document_details = {'id': worker_document.id, 'name': worker_document.name,
                                       'expiry_date': worker_document.expiry_date, 'done': worker_document.done,
                                       'file_url': get_file_url()}
            worker_documents.append(worker_document_details)
        worker_details['worker_documents'] = worker_documents
    work_times = WorkTime.objects.filter(worker=worker, date__month=datetime.today().month).order_by('date')
    work_times_response = []
    if work_times.exists():
        for work_time in work_times:
            work_time_details = {'id': work_time.id, 'cost': work_time.cost, 'start': work_time.start,
                                 'pause': work_time.pause, 'end': work_time.end, 'duration': work_time.duration,
                                 'date': work_time.date}
            work_times_response.append(work_time_details)
    worker_details['worktimes'] = work_times_response
    return worker_details


@api.get('/get-worktimes', tags=['Worker'])
def get_worktimes(request, worker: int = Query(...), month: str = Query(...), year: str = Query(...)):
    worker = get_object_or_404(Worker, id=worker)
    work_times = WorkTime.objects.filter(worker=worker, date__month=month, date__year=year).order_by('date')
    work_times_response = []
    if work_times.exists():
        for work_time in work_times:
            work_time_details = {'id': work_time.id, 'cost': work_time.cost, 'start': work_time.start,
                                 'pause': work_time.pause, 'end': work_time.end, 'duration': work_time.duration,
                                 'status': work_time.status, 'date': work_time.date}
            work_times_response.append(work_time_details)
    return work_times_response


@api.post('/update-worker', tags=['Worker'])
def update_worker(request, data: UpdateWorker = Form(...)):
    worker = get_object_or_404(Worker, id=data.worker_id)
    log_input_string = ''
    if data.new_worker_id:
        if Worker.objects.filter(worker_id=data.new_worker_id).exists():
            return HttpResponse('Es existiert bereits ein Mitarbeiter mit dieser Personalnummer', status=406)
        log_input_string += f'Personalnummer: {worker.worker_id} => {data.new_worker_id}'
        worker.worker_id = data.new_worker_id
    if data.name:
        log_input_string += f'Name: {worker.name} => {data.name} ; '
        worker.name = data.name
    if data.position:
        def get_position():
            try:
                position = get_object_or_404(Position, name=data.position)
            except Http404:
                position = Position.objects.create(name=data.position)
            return position

        log_input_string += f'Position: {worker.position} => {data.position} ; '
        worker.position = get_position()
    if worker.position.name == 'Fahrer':
        worker.is_driver = True
    else:
        worker.is_driver = False
    if data.start_date:
        worker.start_date = data.start_date
        log_input_string += f'Arbeitsbeginn: {datetime.strftime(datetime.strptime(worker.start_date, "%Y-%m-%d"), "%d.%m.%Y")} => {datetime.strftime(datetime.strptime(data.start_date, "%Y-%m-%d"), "%d.%m.%Y")} ; '
    if data.end_date:
        log_input_string += f'Kündigung: {datetime.strftime(datetime.strptime(data.end_date, "%Y-%m-%d"), "%d.%m.%Y")}'
        worker.has_quit = True
        worker.quit_date = data.end_date
    if data.salary:
        log_input_string += f'Lohn: {worker.salary} => {data.salary} ; '
        worker.salary = data.salary
    if data.daily_expense:
        log_input_string += f'Spesen: {worker.daily_expense} => {data.daily_expense} ; '
        worker.daily_expense = data.daily_expense
    if data.holidays:
        log_input_string += f'Urlaubstage: {worker.holidays} => {data.holidays} ; '
        worker.holidays = data.holidays
    if data.note:
        log_input_string += f'Notizen: {worker.note} => {data.note} ; '
        worker.note = data.note
    if data.has_quit and data.quit_date:
        log_input_string += f'Kündigung: {data.quit_date} ; '
        worker.has_quit = True
        worker.quit_date = data.quit_date
    if not data.quit_date:
        log_input_string += f'Erneuter Einstieg: {today.strftime("%d.%m.%Y")} ; '
        worker.has_quit = False
        worker.quit_date = None
    if data.remaining_holidays:
        if HolidayAccount.objects.filter(worker=worker, year=today.year).exists():
            ha = HolidayAccount.objects.get(worker=worker, year=today.year)
            ha.remaining_holiday_days = int(data.remaining_holidays)
            ha.save()
        else:
            HolidayAccount.objects.create(worker=worker, year=today.year,
                                          remaining_holiday_days=int(data.remaining_holidays))
    worker.save()
    log_input = f'Details zum Mitarbeiter {worker.name} wurden verändert. Veränderungen: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=worker.own_firm,
                       log_input=log_input)
    return 200


@api.post('/delete-worker', tags=['Worker'])
def delete_worker(request, worker_id: int = Form(...), admin: str = Form(...)):
    worker = get_object_or_404(Worker, id=worker_id)
    worker.delete()
    log_input = f'Mitarbeiter {worker.name}({worker.worker_id}) wurde gelöscht'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=worker.own_firm,
                       log_input=log_input)
    return 200


@api.post('/add-debt', tags=['Worker'])
def add_debt(request, data: AddDebt = Form(...)):
    worker = get_object_or_404(Worker, id=data.worker_id)
    DebtPayment.objects.create(worker=worker, amount=Decimal(data.amount), date=data.date, notes=data.notes)
    if data.amount >= 0:
        log_input = f'Schulden in Wert von {data.amount} € wurde an {worker.name}({worker.worker_id}) gegeben.'
    else:
        log_input = f'{worker.name} mit der Personalnummer {worker.worker_id} hat {data.amount} € an Schulden abgeglichen.'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=worker.own_firm,
                       log_input=log_input)
    return 200


@api.post('/update-debt', tags=['Worker'])
def update_debt(request, data: UpdateDebt = Form(...)):
    debt = get_object_or_404(DebtPayment, id=data.debt_id)
    log_input_string = ''
    if data.amount:
        log_input_string += f'Betrag: {debt.amount} => {data.amount} ; '
        debt.amount = data.amount
    if data.date:
        log_input_string += f'Datum: {debt.date} => {data.date} ; '
        debt.date = data.date
    if data.notes:
        log_input_string += f'Notizen: {debt.notes} => {data.notes} ; '
        debt.notes = data.notes
    debt.save()
    log_input = f'Details zur Schulden vom Mitarbeiter {debt.worker.name} wurden verändert. Veränderungen: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=debt.worker.own_firm,
                       log_input=log_input)
    return 200


@api.post('/delete-debt', tags=['Worker'])
def delete_debt(request, debt_id: int = Form(...)):
    debt = get_object_or_404(DebtPayment, id=debt_id)
    debt.delete()
    return 200


@api.post('/add-worker-document', tags=['Worker'])
def add_worker_document(request, data: AddWorkerDocument = Form(...)):
    worker = get_object_or_404(Worker, id=data.worker_id)
    document = WorkerDocument.objects.create(worker=worker, name=data.name, expiry_date=data.expiry_date)
    if data.done:
        document.done = True
        document.save()
    log_input = f'Neues Dokument {document.name} wurde zum Mitarbeiter {worker.name}({worker.worker_id}) hinzugefügt'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=worker.own_firm,
                       log_input=log_input)
    return document.id


@api.post('/add-worker-document-file', tags=['Worker'])
def add_worker_document_file(request, document_id: int = Form(...), file: UploadedFile = File(...)):
    document = get_object_or_404(WorkerDocument, id=document_id)
    document.file = file
    document.save()
    return 200


@api.post('/update-worker-document', tags=['Worker'])
def update_worker_document(request, data: UpdateWorkerDocument = Form(...)):
    document = get_object_or_404(WorkerDocument, id=data.document_id)
    log_input_string = ''
    if data.name:
        log_input_string += f'Name: {document.name} => {data.name}; '
        document.name = data.name
    if data.expiry_date:
        log_input_string += f'Verfallsdatum: {document.expiry_date} => {data.expiry_date}; '
        document.expiry_date = data.expiry_date
    if data.delete_file:
        log_input_string += f'Dokument gelöscht: {document.file} '
        document.file.delete()
    if data.done:
        log_input_string += f'Erledigt ✅; '
        document.done = True
    if data.done is False:
        log_input_string += f'Nicht Erledigt ❌; '
        document.done = False
    document.save()
    log_input = f'Details vom Mitarbeiter-Dokument/Termin {document.name} wurden verändert. Veränderungen: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=document.worker.own_firm,
                       log_input=log_input)
    return 200


@api.post('/change-daily-expenses', tags=['Worker'])
def change_daily_expenses(request, worker_id: int = Form(...), admin: str = Form(...), amount: float = Form(...)):
    worker = get_object_or_404(Worker, id=worker_id)
    worker_amount = worker.daily_expense
    worker.daily_expense = amount
    worker.save()
    log_input = f'Spesenbetrag des Mitarbeiters {worker.name}({worker.worker_id}) wurde von {worker_amount} € auf {amount} € verändert.'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=worker.own_firm,
                       log_input=log_input)
    return 200


@api.post('/add-worktime', tags=['Worker'])
def add_worktime(request, data: WorkTimeSchema = Form(...)):
    worker = get_object_or_404(Worker, id=data.worker_id)
    if WorkTime.objects.filter(worker=worker, date=data.date).exists():
        worktime = WorkTime.objects.filter(worker=worker, date=data.date).last()
    else:
        worktime = WorkTime.objects.create(worker=worker, date=data.date)
    worktime.start = data.start
    worktime.pause = data.pause
    worktime.end = data.end
    worktime.duration = data.duration
    if data.daily_expenses:
        worktime.cost = data.daily_expenses
    else:
        worktime.cost = worker.daily_expense
    worktime.save()
    log_input = f'Arbeitszeiten für den Mitarbeiter {worker.name}({worker.worker_id}) wurde für den {datetime.strftime(datetime.strptime(data.date, "%Y-%m-%d"), "%d.%m.%Y")} eingefügt'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=worker.own_firm,
                       log_input=log_input)
    return 200


@api.post('/delete-worktime', tags=['Worker'])
def delete_worktime(request, worktime_id: int = Form(...)):
    worktime = get_object_or_404(WorkTime, id=worktime_id)
    worktime.delete()
    return 200


@api.get('/get-daily-expenses', tags=['Worker'], description='order_by options: name, worker_id, daily_expense')
def get_daily_expenses(request, query: GetDailyExpenses = Query(...)):
    if query.order_by:
        if query.order_by == 'name':
            if query.direction == 'asc':
                qs = Worker.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('name')
            elif query.direction == 'des':
                qs = Worker.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-name')
        elif query.order_by == 'worker_id':
            if query.direction == 'asc':
                qs = Worker.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'worker_id')
            elif query.direction == 'des':
                qs = Worker.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-worker_id')
        elif query.order_by == 'daily_expense':
            if query.direction == 'asc':
                qs = Worker.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    'daily_expense')
            elif query.direction == 'des':
                qs = Worker.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                    '-daily_expense')
    else:
        qs = Worker.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm))
    response = []
    for worker in qs:
        worker_details = {'id': worker.id, 'name': worker.name, 'worker_id': worker.worker_id,
                          'daily_expense': worker.daily_expense}
        response.append(worker_details)
    return response


# OFFDAYS

@api.get('/get-offdays', tags=['Offdays'])
def get_offdays(request, query: GetOffdays = Query(...)):
    own_firm = get_object_or_404(OwnFirm, name=query.own_firm)
    response = []
    for worker in Worker.objects.filter(own_firm=own_firm, is_working=True):
        worker_offday_details = {'name': worker.name, 'worker_id': worker.id}
        qs = Offday.objects.filter(worker=worker, date__month=query.month, date__year=query.year).order_by('date')
        if query.holidays_remaining_start and query.holidays_remaining_end:
            if HolidayAccount.objects.filter(worker=worker, year=datetime.now().year).exists():
                if not int(query.holidays_remaining_start) <= \
                       HolidayAccount.objects.filter(worker=worker, year=datetime.now().year)[
                           0].remaining_holiday_days <= int(query.holidays_remaining_end):
                    qs = qs.exclude(worker__id=worker.id)
        if query.worker:
            if worker == get_object_or_404(Worker, id=int(query.worker)):
                qs = Offday.objects.filter(worker=worker, date__month=query.month, date__year=query.year).order_by(
                    'date')
            else:
                qs = []
        if query.tag:
            qs = Offday.objects.filter(worker=worker, date__month=query.month, date__year=query.year,
                                       tag__id=int(query.tag)).order_by('date')
        offdays_list = []
        if qs:
            for offday in qs:
                offdays_details = {'id': offday.id, 'date': offday.date, 'tag': offday.tag.name,
                                   'tag_colour': offday.tag.colour,
                                   'notes': offday.notes}
                offdays_list.append(offdays_details)
        worker_offday_details['offdays'] = offdays_list
        response.append(worker_offday_details)
    return response


@api.post('/create-offdays', tags=['Offdays'])
def create_offdays(request, data: CreateOffdays = Form(...)):
    worker = get_object_or_404(Worker, id=data.worker_id)
    dates_list = [i.split(',') for i in data.dates]
    log_input_string = ''
    for date in dates_list[0]:
        date_object = datetime.strptime(date, '%Y-%m-%d')
        if Offday.objects.filter(worker=worker, date=date).exists():
            for offday in Offday.objects.filter(worker=worker, date=date):
                offday.delete()
        if date_object.weekday() < 5:
            log_input_string += f''
            created_offday = Offday.objects.create(worker=worker, tag=get_object_or_404(OffdayTag, name=data.tag),
                                                   date=date, notes=data.notes)
            if created_offday.tag.name == 'Urlaub':
                if HolidayAccount.objects.filter(worker=worker, year=date_object.year):
                    holiday_account = HolidayAccount.objects.get(worker=worker, year=date_object.year)
                    holiday_account.remaining_holiday_days -= 1
                    holiday_account.used_holiday_days += 1
                    holiday_account.save()
                else:
                    HolidayAccount.objects.create(worker=worker, used_holiday_days=1,
                                                  remaining_holiday_days=worker.holidays - 1, year=date_object.year)
            log_input_string += f'{date_object.strftime("%d.%m.%Y")} - {created_offday.tag.name} ; '
    log_input = f'Urlaubsplanung für Mitarbeiter {worker.name}({worker.worker_id}) hinzugefügt: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=data.admin), own_firm=worker.own_firm,
                       log_input=log_input)
    return 200


@api.post('/delete-offdays', tags=['Offdays'])
def delete_offdays(request, data: List[str] = Form(...), admin: str = Form(...)):
    offdays_list = [i.split(',') for i in data]
    log_input_string = ''
    own_firm = None
    worker = None
    n = 0
    for offday in offdays_list[0]:
        qs = Offday.objects.get(id=int(offday))
        if n == 0:
            worker = qs.worker
            own_firm = worker.own_firm
            n += 1
        log_input_string += f'{qs.date.strftime("%d.%m.%Y")} : {qs.tag.name} ; '
        qs.delete()
    log_input = f'Tage von der Urlaubsplanung für {worker.name}({worker.worker_id}) wurden gelöscht: {log_input_string}'
    Log.objects.create(admin=get_object_or_404(HarbyAdmin, user_hash=admin), own_firm=own_firm,
                       log_input=log_input)
    return 200


@api.get('/get-urlaubskonto', tags=['Urlaubskonto'])
def get_urlaubskonto(request, own_firm: str = Query(...), year: str = Query(...)):
    def own_firm_query():
        own_firm_data = get_object_or_404(OwnFirm, name=own_firm)
        return own_firm_data

    workers = Worker.objects.filter(own_firm=own_firm_query(), is_working=True)
    months = ['jan', 'feb', 'mär', 'apr', 'mai', 'jun', 'jul', 'aug', 'sep', 'okt', 'nov', 'dez']
    response = {}

    md = HolidayAccount.objects.all().values('year').distinct()
    available_years = []
    for i in md:
        available_years.append(i['year'])

    response['available_years'] = available_years

    details = []
    for worker in workers:
        worker_holiday_details = {'name': worker.name, 'days': {}}
        for month in months:
            worker_holiday_details['days'][month] = {'Urlaub': [], 'Krank': [], 'Unbezahlter Urlaub': []}
        if HolidayAccount.objects.filter(worker=worker, year=year).exists():
            holiday_account = HolidayAccount.objects.get(worker=worker, year=year)
            worker_holiday_details['remaining_holiday_days'] = holiday_account.remaining_holiday_days
            worker_holiday_details['used_holiday_days'] = holiday_account.used_holiday_days
        else:
            worker_holiday_details['remaining_holiday_days'] = worker.holidays
            worker_holiday_details['used_holiday_days'] = 0
        for holiday in Offday.objects.filter(worker=worker, date__year=year):
            if holiday.tag.name == 'Urlaub':
                worker_holiday_details['days'][months[holiday.date.month - 1]]['Urlaub'].append(holiday.date)
            elif holiday.tag.name == 'Krank':
                worker_holiday_details['days'][months[holiday.date.month - 1]]['Krank'].append(holiday.date)
            elif holiday.tag.name == 'Unbezahlter Urlaub':
                worker_holiday_details['days'][months[holiday.date.month - 1]]['Unbezahlter Urlaub'].append(
                    holiday.date)
        details.append(worker_holiday_details)
    response['details'] = details
    return response


@api.get('/get-tags', tags=['Settings'])
def get_tags(request):
    tour_tags = TourStatus.objects.all()
    contacts_tags = ContactTag.objects.all()
    offdays_tags = OffdayTag.objects.all()
    tour = []
    contacts = []
    offdays = []
    for tour_tag in tour_tags:
        tour.append({'id': tour_tag.id, 'name': tour_tag.name, 'colour': tour_tag.colour})
    for contacts_tag in contacts_tags:
        contacts.append({'id': contacts_tag.id, 'name': contacts_tag.name, 'colour': contacts_tag.colour})
    for offdays_tag in offdays_tags:
        offdays.append({'id': offdays_tag.id, 'name': offdays_tag.name, 'colour': offdays_tag.colour})
    response = {'tour_tags': tour, 'contacts_tags': contacts, 'offday_tags': offdays}
    return response


@api.get('/get-available-colours', tags=['Settings'])
def get_available_colours(request):
    response = []
    for colour in Colour.objects.all():
        colour_details = {'id': colour.id, 'hex_code': colour.colour_hex}
        response.append(colour_details)
    return response


@api.post('/create-tag', tags=['Settings'])
def create_tag(request, data: CreateTag = Form(...)):
    if data.section == 'tour':
        try:
            TourStatus.objects.create(name=data.name, colour=data.colour)
        except IntegrityError:
            return HttpResponse('Dieses Etikett für die Tourplanung existiert beretis.', status=406)
    if data.section == 'contacts':
        try:
            ContactTag.objects.create(name=data.name, colour=data.colour)
        except IntegrityError:
            return HttpResponse('Dieses Etikett für die Kontaktendatenbank existiert beretis.', status=406)
    if data.section == 'offday':
        try:
            OffdayTag.objects.create(name=data.name, colour=data.colour)
        except IntegrityError:
            return HttpResponse('Dieses Etikett für die Urlaubsplanung existiert beretis.', status=406)
    return 200


@api.post('/update-tag', tags=['Settings'])
def update_tag(request, data: UpdateTag = Form(...)):
    if data.section == 'tour':
        tag = get_object_or_404(TourStatus, id=data.id)
        if data.name:
            tag.name = data.name
        if data.colour:
            tag.colour = data.colour
        tag.save()
    if data.section == 'contacts':
        tag = get_object_or_404(ContactTag, id=data.id)
        if data.name:
            tag.name = data.name
        if data.colour:
            tag.colour = data.colour
        tag.save()
    if data.section == 'offday':
        tag = get_object_or_404(OffdayTag, id=data.id)
        if data.name:
            tag.name = data.name
        if data.colour:
            tag.colour = data.colour
        tag.save()
    return 200


@api.post('/delete-tag', tags=['Settings'])
def delete_tag(request, data: DeleteTag = Form(...)):
    if data.section == 'tour':
        tag = get_object_or_404(TourStatus, id=data.id)
        tag.delete()
    if data.section == 'contacts':
        tag = get_object_or_404(ContactTag, id=data.id)
        tag.delete()
    if data.section == 'offday':
        tag = get_object_or_404(OffdayTag, id=data.id)
        tag.delete()
    return 200


@api.get('/get-customers', tags=['Settings'], description='order_by options: name, vat, address')
def get_customers(request, query: GetCustomers = Query(...)):
    response = []
    if query.order_by:
        if query.order_by == 'name':
            if query.direction == 'asc':
                qs = Firm.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('name')
            elif query.direction == 'des':
                qs = Firm.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-name')
        elif query.order_by == 'vat':
            if query.direction == 'asc':
                qs = Firm.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('vat')
            elif query.direction == 'des':
                qs = Firm.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-vat')
        elif query.order_by == 'address':
            if query.direction == 'asc':
                qs = Firm.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('address')
            elif query.direction == 'des':
                qs = Firm.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-address')
    else:
        qs = Firm.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-id')
    for firm in qs:
        firm_details = {'id': firm.id, 'firm_name': firm.name, 'vat': firm.vat, 'address': firm.address}
        response.append(firm_details)

    return response


@api.post('/create-customer', tags=['Settings'])
def create_customer(request, data: CreateCustomer = Form(...)):
    def own_firm_query():
        own_firm_data = get_object_or_404(OwnFirm, name=data.own_firm)
        return own_firm_data

    try:
        Firm.objects.create(own_firm=own_firm_query(), name=data.firm_name, vat=data.vat, address=data.address)
    except IntegrityError:
        return HttpResponse('Es existiert bereits eine Firma mit diesen Namen', status=406)
    return 200


@api.post('/update-customer', tags=['Settings'])
def update_customer(request, data: UpdateCustomer = Form(...)):
    firm = get_object_or_404(Firm, id=data.firm_id)
    if data.name:
        if Firm.objects.filter(name=data.name).exists():
            return HttpResponse('Es existiert bereits eine Firma mit diesen Namen', status=406)
        firm.name = data.name
    if data.vat:
        firm.vat = data.vat
    if data.address:
        firm.address = data.address
    firm.save()
    return 200


@api.post('/delete-customer', tags=['Settings'])
def delete_customer(request, data: DeleteCustomer = Form(...)):
    firm = get_object_or_404(Firm, id=data.firm_id)
    firm.delete()
    return 200


@api.post('/create-admin', tags=['Settings'])
def create_admin(request, data: CreateAdmin = Form(...)):
    own_firms = []
    own_firms_data = [i.split(',') for i in data.own_firms]
    for own_firm in own_firms_data[0]:
        own_firms.append(get_object_or_404(OwnFirm, name=own_firm))

    try:
        validate_password(data.password)
    except ValidationError:
        return HttpResponse(
            'Passwort nicht gültig. Ein Passwort muss mindestens 8 Buchstaben, Zahlen oder Zeichen beinhalten, darf nicht deinen Persönlichen Daten ähneln, nicht komplett aus Ziffern entstehen und darf nicht ein häufig benutztes Password sein wie "passwort" oder "12345678".',
            status=406)

    if User.objects.filter(username=data.username).exists():
        return HttpResponse('Dieser Benutzername wird bereits benutzt.', status=406)

    user = User.objects.create(username=data.username, first_name=data.name, password=data.password)

    def uuid4_generator():
        test = uuid.uuid4()
        while HarbyAdmin.objects.filter(user_hash=test).exists():
            test = uuid.uuid4()
        return test

    harby_admin = HarbyAdmin.objects.create(user=user, user_hash=uuid4_generator())
    for own_firm in own_firms:
        harby_admin.own_firms.add(own_firm)
    return 200


@api.get('/get-admin-details', tags=['Settings'])
def get_admin_details(request, data: GetAdminDetails = Query(...)):
    admin = get_object_or_404(HarbyAdmin, user__id=data.admin_id)
    user = admin.user
    own_firms = []
    all_own_firms = []
    for own_firm in OwnFirm.objects.all():
        all_own_firms.append({'id': own_firm.id, 'name': own_firm.name})
    for own_firm in admin.own_firms.all():
        own_firms.append({'id': own_firm.id, 'name': own_firm.name})
    response = {'name': user.first_name, 'username': user.username, 'password': user.password,
                'all_own_firms': all_own_firms, 'access_to_own_firms': own_firms}
    return response


@api.post('/update-admin', tags=['Settings'])
def update_admin(request, data: UpdateAdmin = Form(...)):
    user = get_object_or_404(User, id=data.id)
    admin = get_object_or_404(HarbyAdmin, user=user)
    if data.name:
        user.first_name = data.name
    if data.username:
        if User.objects.filter(username=data.username).exists():
            return HttpResponse('Dieser Benutzername wird bereits benutzt.', status=406)
        user.username = data.username
    if data.password:
        try:
            validate_password(data.password)
        except ValidationError:
            return HttpResponse(
                'Passwort nicht gültig. Ein Passwort muss mindestens 8 Buchstaben, Zahlen oder Zeichen beinhalten, darf nicht deinen Persönlichen Daten ähneln, nicht komplett aus Ziffern entstehen und darf nicht ein häufig benutztes Password sein wie "passwort" oder "12345678".',
                status=406)
        user.password = data.password
    if data.own_firms:
        admin.own_firms.clear()
        own_firms_data = [i.split(',') for i in data.own_firms]
        for own_firm in own_firms_data[0]:
            admin.own_firms.add(get_object_or_404(OwnFirm, name=own_firm))
    user.save()
    return 200


@api.post('/delete-admin', tags=['Settings'])
def delete_admin(request, admin_id: int = Form(...)):
    user = get_object_or_404(User, id=admin_id)
    user.delete()
    return 200


@api.get('/get-admins', tags=['Settings'], description='order_by options: name, username')
def get_admins(request, query: GetAdmins = Query(...)):
    response = []
    if query.order_by:
        if query.order_by == 'name':
            if query.direction == 'asc':
                qs = HarbyAdmin.objects.all().order_by('user__first_name')
            elif query.direction == 'des':
                qs = HarbyAdmin.objects.all().order_by('-user__first_name')
        elif query.order_by == 'username':
            if query.direction == 'asc':
                qs = HarbyAdmin.objects.all().order_by('user__username')
            elif query.direction == 'des':
                qs = HarbyAdmin.objects.all().order_by('-user__username')
    else:
        qs = HarbyAdmin.objects.all()
    for admin in qs:
        admin_details = {'id': admin.user.id, 'name': admin.user.first_name, 'username': admin.user.username}
        response.append(admin_details)
    return response


@api.get('/get-logs', tags=['Settings'], description='order_by options: admin, date, firm')
def get_logs(request, query: GetLogs = Query(...)):
    response = []

    def log_query():
        if query.order_by:
            if query.order_by == 'admin':
                if query.direction == 'asc':
                    qs = Log.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                        'admin__user__first_name')
                elif query.direction == 'des':
                    qs = Log.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                        '-admin__user__first_name')
            elif query.order_by == 'date':
                if query.direction == 'asc':
                    qs = Log.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                        'timestamp')
                elif query.direction == 'des':
                    qs = Log.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                        '-timestamp')
            elif query.order_by == 'firm':
                if query.direction == 'asc':
                    qs = Log.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                        'own_firm__name')
                elif query.direction == 'des':
                    qs = Log.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by(
                        '-own_firm__name')
        else:
            qs = Log.objects.filter(own_firm=get_object_or_404(OwnFirm, name=query.own_firm)).order_by('-id')
        if query.admin:
            qs = qs.filter(admin=get_object_or_404(HarbyAdmin, user__id=query.admin))
        if query.start_date:
            qs = qs.filter(timestamp__gte=query.start_date)
        if query.end_date:
            qs = qs.filter(timestamp__lte=query.end_date)
        return qs

    def get_admin(log_obj):
        if log.admin:
            return log.admin.user.first_name
        else:
            return None

    for log in log_query():
        log_details = {'admin': get_admin(log), 'date': log.timestamp.strftime('%d.%m.%Y %H:%M:%S'),
                       'own_firm': log.own_firm.name, 'event': log.log_input}
        response.append(log_details)

    return response


@api.get('/get-own-firms', tags=['Own Firm'])
def get_own_firms(request, query: GetOwnFirm = Query(...)):
    if query.admin:
        admin = get_object_or_404(HarbyAdmin, user_hash=query.admin)
        response = []
        for own_firm in admin.own_firms.all():
            def logo():
                if own_firm.logo:
                    return root + own_firm.logo.url
                else:
                    return None

            own_firm_details = {'id': own_firm.id, 'name': own_firm.name, 'logo': logo(),
                                'address': own_firm.address, 'phone': own_firm.phone, 'mail': own_firm.mail,
                                'chairman': own_firm.chairman,
                                'company_place': own_firm.company_place, 'register_court': own_firm.register_court,
                                'tax_nr': own_firm.tax_nr, 'ustid': own_firm.ustid,
                                'contact_name': own_firm.contact_name, 'contact_phone': own_firm.contact_phone,
                                'contact_fax': own_firm.contact_fax, 'bank_name': own_firm.bank_name,
                                'bank_iban': own_firm.bank_iban, 'bank_bic': own_firm.bank_bic}
            response.append(own_firm_details)
        return response
    else:
        response = []
        for own_firm in OwnFirm.objects.all():
            def logo():
                if own_firm.logo:
                    return root + own_firm.logo.url
                else:
                    return None

            own_firm_details = {'id': own_firm.id, 'name': own_firm.name, 'logo': logo(),
                                'address': own_firm.address, 'phone': own_firm.phone, 'mail': own_firm.mail,
                                'chairman': own_firm.chairman,
                                'company_place': own_firm.company_place, 'register_court': own_firm.register_court,
                                'tax_nr': own_firm.tax_nr, 'ustid': own_firm.ustid,
                                'contact_name': own_firm.contact_name, 'contact_phone': own_firm.contact_phone,
                                'contact_fax': own_firm.contact_fax, 'bank_name': own_firm.bank_name,
                                'bank_iban': own_firm.bank_iban, 'bank_bic': own_firm.bank_bic}
            response.append(own_firm_details)
        return response


@api.post('/create-own-firm', tags=['Own Firm'])
def create_own_firm(request, data: CreateOwnFirm = Form(...)):
    try:
        created_own_firm = OwnFirm.objects.create(name=data.name, address=data.address, phone=data.phone,
                                                  mail=data.mail,
                                                  chairman=data.chairman, company_place=data.company_place,
                                                  register_court=data.register_court,
                                                  tax_nr=data.tax_nr, ustid=data.ustid,
                                                  contact_name=data.contact_name,
                                                  contact_phone=data.contact_phone, contact_fax=data.contact_fax,
                                                  bank_name=data.bank_name,
                                                  bank_iban=data.bank_iban, bank_bic=data.bank_bic)
    except IntegrityError:
        return HttpResponse('Du hast bereits eine Firma mit diesen Namen.', status=406)
    return created_own_firm.id


@api.post('/own-firm-add-logo', tags=['Own Firm'])
def own_firm_add_logo(request, own_firm_id: int = Form(...), file: UploadedFile = File(...)):
    own_firm = get_object_or_404(OwnFirm, id=own_firm_id)
    own_firm.logo = file
    own_firm.save()
    return 200


@api.post('/update-own-firm', tags=['Own Firm'])
def update_own_firm(request, data: UpdateOwnFirm = Form(...)):
    own_firm = get_object_or_404(OwnFirm, id=data.id)
    if data.name:
        if OwnFirm.objects.filter(name=data.name).exists():
            return HttpResponse('Du hast bereits eine Firma mit diesen Namen.', status=406)
        own_firm.name = data.name
    if data.address:
        own_firm.address = data.address
    if data.phone:
        own_firm.phone = data.phone
    if data.mail:
        own_firm.mail = data.mail
    if data.chairman:
        own_firm.chairman = data.chairman
    if data.company_place:
        own_firm.company_place = data.company_place
    if data.register_court:
        own_firm.register_court = data.register_court
    if data.tax_nr:
        own_firm.tax_nr = data.tax_nr
    if data.ustid:
        own_firm.ustid = data.ustid
    if data.contact_name:
        own_firm.contact_name = data.contact_name
    if data.contact_phone:
        own_firm.contact_phone = data.contact_phone
    if data.contact_fax:
        own_firm.contact_fax = data.contact_fax
    if data.bank_name:
        own_firm.bank_name = data.bank_name
    if data.bank_iban:
        own_firm.bank_iban = data.bank_iban
    if data.bank_bic:
        own_firm.bank_bic = data.bank_bic
    own_firm.save()
    return 200


@api.post('/delete-own-firm', tags=['Own Firm'])
def delete_own_firm(request, own_firm_id: int = Form(...)):
    own_firm = get_object_or_404(OwnFirm, id=own_firm_id)
    own_firm.delete()
    return 200


@api.post('/login', tags=['Authentication'])
def login(request, data: Login = Form(...)):
    user = get_object_or_404(User, username=data.username)
    response = {'hash': HarbyAdmin.objects.get(user=user).user_hash, 'name': user.first_name}
    if user.is_superuser:
        response['role'] = 'superuser'
    else:
        response['role'] = 'user'
    return response
